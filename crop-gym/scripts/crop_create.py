import os
from pcse_gym.envs.common_env import PCSEEnv
import pcse_gym.envs.common_env as _common_env_module
import yaml
from datetime import timedelta, date
import requests
import math
import openpyxl
from pcse.fileinput import ExcelWeatherDataProvider

_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
_WEATHER_INPUT  = os.path.join(_SCRIPTS_DIR, 'weather', 'dummy_file.xlsx')
_WEATHER_OUTPUT = os.path.join(_SCRIPTS_DIR, 'weather', 'data_weather.xlsx')


class CreateCrop(PCSEEnv):

    def __init__(self,
                 latitude: float,
                 longitude: float,
                 *args,
                 **kwargs):

        self.latitude = latitude
        self.longitude = longitude
        self.args = args
        self.kwargs = kwargs
        self.get_range_date()

        # Fetches weather from Open-Meteo and writes to _WEATHER_OUTPUT.
        # self._excel_output_path is set here so the patched provider can use it.
        self._fetch_enriched_weather()

        location = (latitude, longitude)

        # Temporarily replace the module-level NASA Power factory so the parent
        # __init__ uses our Excel file instead of making a network request.
        _orig = _common_env_module.get_weather_data_provider
        _common_env_module.get_weather_data_provider = (
            lambda _loc: ExcelWeatherDataProvider(self._excel_output_path)
        )
        try:
            super().__init__(location=location, *args, **kwargs)
        finally:
            _common_env_module.get_weather_data_provider = _orig

        self.crop_data = []
        self.days_gone = 0

    def aply_action(self, action: dict):
        """
        Apply a management action to the crop and advance 1 day in simulation.

        action keys:
          'irrigation': water amount in mm
          'N': nitrogen fertilizer amount
        """
        prev_obs = self.crop_data[-1]["obs"] if self.crop_data else None

        obs, reward, done, truncated, info = self.step(action)

        self.crop_data.append({
            "prev_obs":  prev_obs,
            "obs":       obs,
            "reward":    reward,
            "done":      done,
            "truncated": truncated,
            "info":      info,
            "days_gone": self.days_gone,
            "action":    action,
        })

        self.days_gone += 1

    def _fetch_enriched_weather(self):
        """
        Fetches daily weather from the Open-Meteo archive API, converts it to
        the format expected by PCSE, and writes it to data_weather.xlsx.
        Sets self._excel_output_path for use by _init_pcse_model.
        """
        self._excel_output_path = _WEATHER_OUTPUT

        print(f"Buscando dados da Open-Meteo para Lat={self.latitude}, Lon={self.longitude}")
        base_url = "https://archive-api.open-meteo.com/v1/archive"

        params_pcse = {
            "latitude":  self.latitude,
            "longitude": self.longitude,
            "start_date": self.crop_start_date.strftime("%Y-%m-%d"),
            "end_date":   self.crop_end_date.strftime("%Y-%m-%d"),
            "daily": [
                "temperature_2m_max", "temperature_2m_min", "precipitation_sum",
                "relative_humidity_2m_mean", "wind_speed_10m_max", "shortwave_radiation_sum",
            ],
            "timezone": "auto",
            "wind_speed_unit": "ms",
            "precipitation_unit": "mm",
        }

        try:
            print("Executando chamada à API Open-Meteo...")
            response_pcse = requests.get(base_url, params=params_pcse)
            response_pcse.raise_for_status()
            data_pcse = response_pcse.json()
        except requests.exceptions.RequestException as e:
            print(f"Erro ao buscar dados da Open-Meteo: {e}")
            raise

        self.elevation = data_pcse.get('elevation', 0.0)
        print(f"Elevação do local: {self.elevation}m")

        daily_data = data_pcse['daily']
        time_list  = daily_data['time']

        self.enriched_weather_data = []

        for i, day_str in enumerate(time_list):
            try:
                day  = date.fromisoformat(day_str)
                tmin = float(daily_data['temperature_2m_min'][i])
                tmax = float(daily_data['temperature_2m_max'][i])
                rh   = float(daily_data['relative_humidity_2m_mean'][i])

                # Vapour pressure (kPa) via Tetens equation
                es_tmin = 0.6108 * math.exp((17.27 * tmin) / (tmin + 237.3))
                es_tmax = 0.6108 * math.exp((17.27 * tmax) / (tmax + 237.3))
                vap = (rh / 100.0) * ((es_tmin + es_tmax) / 2.0)

                # Radiation: J/m² → kJ/m²
                irrad_kj = float(daily_data['shortwave_radiation_sum'][i]) / 1000.0

                self.enriched_weather_data.append([
                    day, irrad_kj, tmin, tmax,
                    vap, float(daily_data['wind_speed_10m_max'][i]),
                    float(daily_data['precipitation_sum'][i]), -999,
                ])

            except (ValueError, TypeError, KeyError) as e:
                print(f"Aviso: dados inválidos para {day_str}, pulando. Erro: {e}")
                continue

        if not self.enriched_weather_data:
            raise ValueError("Nenhum dado climático válido foi processado da Open-Meteo.")

        wb = openpyxl.load_workbook(_WEATHER_INPUT)
        ws = wb.active

        # Site metadata (row 9: Longitude, Latitude, Elevation, AngstA, AngstB, HasSunshine)
        ws.cell(row=9, column=1, value=self.longitude)
        ws.cell(row=9, column=2, value=self.latitude)
        ws.cell(row=9, column=3, value=self.elevation)

        # Daily weather rows start at line 13
        for i, row_data in enumerate(self.enriched_weather_data, start=13):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

        wb.save(_WEATHER_OUTPUT)
        print(f"Dados climáticos salvos em {_WEATHER_OUTPUT} ({len(self.enriched_weather_data)} dias)")

    def convert_yml_file(self, file_name):
        with open(file_name) as f:
            return yaml.safe_load(f)[0]

    def get_range_date(self):
        agro_file = self.kwargs.get('agro_config')
        f = self.convert_yml_file(agro_file)
        calendar = f[list(f.keys())[0]]['CropCalendar']
        self.crop_start_date = calendar['crop_start_date']
        self.crop_end_date   = calendar['crop_end_date']
        print(calendar)

    def current_day(self):
        return self.crop_start_date + timedelta(days=self.days_gone)
