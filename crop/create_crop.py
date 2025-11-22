import json
import pandas as pd
import matplotlib.pyplot as plt
from datetime import date, timedelta,datetime
import requests
import math
import openpyxl
from openpyxl.styles import numbers

from pcse.models import Wofost72_WLP_FD
from pcse.fileinput import YAMLAgroManagementReader,CABOFileReader,ExcelWeatherDataProvider
from pcse.util import WOFOST72SiteDataProvider, DummySoilDataProvider

from pcse.base import ParameterProvider, WeatherDataContainer
 


class CropCreation(DummySoilDataProvider):

    def __init__(self,
                 latitude:float, 
                 longitude:float, 
                 start_date:str, 
                 end_date:str, 
                 crop_file:str='soyean_cab.cab', 
                 agro_file:str='agro_file.yaml', 
                 soil_file:str='soil_file.json'):
        
        """
        Busca um conjunto completo de dados climáticos da API Open-Meteo, 
        prontos para serem usados pelo PCSE/WOFOST.

        :param latitude: Latitude do local.
        :param longitude: Longitude do local.
        :param start_date: Data de início str %Y%m%d.
        :param end_date: Data de fim str %Y%m%d.
        :param crop_file: Caminho para o arquivo de parâmetros da cultura (.cab).
        :param agro_file: Caminho para o arquivo de manejo agronômico (.yaml).
        :param soil_file: Caminho para o arquivo de parâmetros do solo (.json).
        """
        
        super().__init__()
        self.wofost = None
        self.output = []
        self.agromanagement = None
        # Renomeado para evitar confusão com self.soil_data (o dicionário)
        self.soil_data_provider = self 
        self.soil_data = None
        self.enriched_weather_data = None
        self.params = None
        self.current_date = None
        self.days_crop = -1 # Iniciar em -1 para que o primeiro dia seja 0

        self.latitude = latitude
        self.longitude = longitude
        self.elevation = None # Adicionado para armazenar a elevação
        self.start_date = datetime.strptime(start_date,'%Y%m%d')
        self.end_date = datetime.strptime(end_date,'%Y%m%d')

        self.crop_file = crop_file
        self.agro_file = agro_file
        self.soil_file = soil_file
        
        self._fetch_enriched_weather()
    
    def _fetch_enriched_weather(self):
        """
        Busca dados climáticos da API Open-Meteo e os formata corretamente para o PCSE.
        """
        print(f"Buscando dados da Open-Meteo para Lat={self.latitude}, Lon={self.longitude}")
        base_url = "https://archive-api.open-meteo.com/v1/archive"
        
        params_pcse = {
            "latitude": self.latitude, "longitude": self.longitude,
            "start_date": self.start_date.strftime("%Y-%m-%d"), "end_date": self.end_date.strftime("%Y-%m-%d"),
            "daily": [
                "temperature_2m_max", "temperature_2m_min", "precipitation_sum",
                "relative_humidity_2m_mean", "wind_speed_10m_max", "shortwave_radiation_sum"
            ],
            # CORREÇÃO: O parâmetro 'elevation' não é válido na URL da API de arquivo.
            # Ele é retornado na resposta por padrão, então o removemos dos parâmetros da requisição.
            "timezone": "auto", "wind_speed_unit": "ms", "precipitation_unit": "mm"
        }
        
        try:
            print("Executando chamada à API Open-Meteo...")
            response_pcse = requests.get(base_url, params=params_pcse)
            response_pcse.raise_for_status()
            data_pcse = response_pcse.json()

        except requests.exceptions.RequestException as e:
            print(f"Erro ao buscar dados da Open-Meteo: {e}")
            raise

        # --- Armazenar Elevação e Processar Dados ---
        self.elevation = data_pcse.get('elevation', 0.0) # Armazena a elevação
        print(f"Elevação do local: {self.elevation}m")
        
        daily_data = data_pcse['daily']
        time_list = daily_data['time']
        
        self.enriched_weather_data = []
        
        for i, day_str in enumerate(time_list):
            try:
                day = date.fromisoformat(day_str)
                tmin = float(daily_data['temperature_2m_min'][i])
                tmax = float(daily_data['temperature_2m_max'][i])
                rh = float(daily_data['relative_humidity_2m_mean'][i])
                
                # Cálculo da pressão de vapor (VAP) em kPa, como esperado pelo PCSE
                es_tmin = 0.6108 * math.exp((17.27 * tmin) / (tmin + 237.3))
                es_tmax = 0.6108 * math.exp((17.27 * tmax) / (tmax + 237.3))
                es_avg = (es_tmin + es_tmax) / 2.0
                vap = (rh / 100.0) * es_avg

                # Converter radiação de J/m2 para kJ/m2
                irrad_kj = float(daily_data['shortwave_radiation_sum'][i]) / 1000.0
                
                row = [
                    day ,irrad_kj, tmin, tmax,
                    vap,float(daily_data['wind_speed_10m_max'][i]),
                    float(daily_data['precipitation_sum'][i]), -999
                    ]
                self.enriched_weather_data.append(row)
            
                
                
            except (ValueError, TypeError, KeyError) as e:
                print(f"Aviso: Dados inválidos ou ausentes para o dia {day_str}. Pulando. Erro: {e}")
                continue
        INPUT_PATH = 'dummy_file.xlsx'
        OUTPUT_PATH = 'data_weather.xlsx'

        START_LINE = 13
        
        wb = openpyxl.load_workbook(INPUT_PATH)
        ws = wb.active    

        if not self.enriched_weather_data:
            raise ValueError("Nenhum dado climático válido foi processado da Open-Meteo.")
        
        for i, data in enumerate(self.enriched_weather_data, start=START_LINE):
            for j, value in enumerate(data, start=1):
                ws.cell(row=i, column=j, value=value)
                # cell = ws.cell(row=i, column=j, value=value)
                # if j == 1 and isinstance(value, (date, datetime)):
                #     cell.number_format = numbers.FORMAT_DATE_MMDDYYYY
        wb.save(OUTPUT_PATH)

    def _days_gone(self):
        self.days_crop += 1

    def initialize_simulation(self):
        """
        Configura e inicializa o modelo de simulação WOFOST.
        """
        print("Inicializando a simulação...")
        
        # Carrega os dados do solo do arquivo JSON
        self.soil_data = self._get_soil_from_json(self.soil_file)
        
        # Transforma a lista de dicionários em um dicionário de listas
        # weather_keywords = {key: [d[key] for d in self.enriched_weather_data] for key in self.enriched_weather_data[0]}

        # print(weather_keywords)
    
        # Inicializa o provedor de dados climáticos, fornecendo TODOS os parâmetros de site necessários
        # weather_data_provider = WeatherDataContainer(**weather_keywords, 
        #                                              location_name=f"Open-Meteo", 
        #                                              LAT=self.latitude,
        #                                              LON=self.longitude,
        #                                              ELEV=self.elevation,
        #                                              ANGSTA=0.1, # Coeficiente de Angstrom A (valor padrão)
        #                                              ANGSTB=0.1) # Coeficiente de Angstrom B (valor padrão)

        weather_data_provider = ExcelWeatherDataProvider('data_weather.xlsx')
        print(f'Keywords {weather_data_provider}')
        
        print('Carregado parametros do clima')
        # Carrega os dados da cultura e do local
        crop_data = CABOFileReader(self.crop_file)
        # crop_data = YAMLCropDataProvider(self.crop_file)
        print('Carregado parametros do soybean')
        site_data = WOFOST72SiteDataProvider(WAV=100)
        
        # Carrega o manejo agronômico
        self.agromanagement = YAMLAgroManagementReader(self.agro_file)
        
        # Passa o dicionário de dados do solo, não a classe
        self.params = ParameterProvider(cropdata=crop_data, soildata=self.soil_data, sitedata=site_data)

        campaigns = list(self.agromanagement)
        print(type(campaigns[0]['CropCalendar']['crop_start_date']))

        
        # Inicializa o modelo WOFOST
        self.wofost = Wofost72_WLP_FD(self.params, weather_data_provider, self.agromanagement)
        self.current_date = self.start_date.date()
        print("Simulação inicializada com sucesso.")

    def _get_soil_from_json(self, file_name):
        print(f"Carregando parâmetros do solo de: {file_name}")
        with open(file_name, 'r') as file:
            content = json.load(file)
        return content

    def irrigation(self, amount_mm: float, efficiency: float = 0.7):
        """
        Agenda um evento de irrigação para a data atual da simulação.
        """
        if not self.wofost:
            print("Erro: A simulação não foi inicializada. Chame initialize_simulation() primeiro.")
            return

        irrigation_event = {
            'date': self.current_date,
            'event_signal': 'apply_irrigation',
            'event_parameters': {
                'amount': amount_mm,
                'efficiency': efficiency
            }
        }
        self.wofost.agromanagement.append(irrigation_event)
        print(f"Evento de irrigação de {amount_mm} mm agendado para {self.current_date}.")

    def _run(self, days: int):
        """
        Executa a simulação por um número especificado de dias.
        """
        if not self.wofost:
            print("Erro: A simulação não foi inicializada. Chame initialize_simulation() primeiro.")
            return
            
        print(f"Executando simulação por {days} dias a partir de {self.current_date}...")
        try:
            for _ in range(days):
                self.wofost.run(days=1)
                # Garante que a saída não esteja vazia antes de adicionar
                current_output = self.wofost.get_output()
                if current_output:
                    self.output.append(current_output[0])
                self.current_date += timedelta(days=1)
                self._days_gone()
        except Exception as e:
            print(f"Simulação encerrada ou encontrou um erro: {e}")

    def plot_simulation(self):
        """
        Plota as principais variáveis de saída da simulação.
        """
        if not self.output:
            print("Nenhum dado de saída para plotar. Execute a simulação primeiro.")
            return
        
        results = pd.DataFrame(self.output)
        results = results.set_index('day')

        fig, axes = plt.subplots(nrows=3, ncols=1, figsize=(12, 15))
        
        results['TAGP'].plot(ax=axes[0], title='Biomassa Total da Parte Aérea (TAGP)', grid=True)
        axes[0].set_ylabel('kg/ha')
        
        results['LAI'].plot(ax=axes[1], title='Índice de Área Foliar (LAI)', grid=True)
        axes[1].set_ylabel('m2/m2')
        
        results['SM'].plot(ax=axes[2], title='Umidade na Zona Radicular (SM)', grid=True)
        axes[2].set_ylabel('cm')

        fig.tight_layout()
        plt.show()

    def get_agent_state(self):
        """
        Monta o vetor de estado completo para o agente de IA.
        """
        if not self.output:
            print("Aviso: Nenhuma saída de simulação para criar o estado.")
            return None
        
        sim_state = self.output[-1] 
        
        sensor_data = {
            'development_stage': round(sim_state['DVS'], 2),
            'biomass_kg_ha': int(sim_state['TAGP']),
            'leaf_area_index': round(sim_state['LAI'], 2),
            'soil_moisture_cm': round(sim_state['SM'], 2),
            'water_stress_factor': round(sim_state['WST'], 2),
            'soil_field_capacity_cm': self.soil_data.get('SMFCF', 0),
            'soil_wilting_point_cm': self.soil_data.get('SMW', 0)
        }

        # Garante que não haja erro de índice se a simulação rodar por mais dias que os dados climáticos
        if self.days_crop < len(self.enriched_weather_data):
            day_data = self.enriched_weather_data[self.days_crop]
        else:
            print("Aviso: Fim dos dados climáticos alcançado.")
            day_data = {} # Retorna dicionário vazio se não houver mais dados

        agent_state = {
            'sensor_data': sensor_data,
            'day_data': day_data
        }
        
        return agent_state

# --- Exemplo de Uso ---
if __name__ == '__main__':
    
    
    # 2. Inicializar a simulação
    latitude_suzano = -23.5434
    longitude_suzano = -46.3106
    simulador = CropCreation(latitude=latitude_suzano,
                             longitude=longitude_suzano,
                             start_date='20230101',
                             end_date='20230401')
                             
    simulador.initialize_simulation()
    
    # # 3. Executar e visualizar
    # simulador._run(days=90)
    # simulador.plot_simulation()

    # # 4. Obter o estado para um agente de IA
    # estado_agente = simulador.get_agent_state()
    # print("\n--- Estado Final para o Agente de IA ---")
    # print(json.dumps(estado_agente, indent=4, default=str))

